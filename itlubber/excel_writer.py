import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
colors = ["2639E9", "F76E6C", "FE7715"]


def save_excel(df, writer, sheet_name, index=False):
    df.to_excel(writer, sheet_name=sheet_name, index=index)
    worksheet = writer.sheets[sheet_name]
    for i, col in enumerate(df.columns, 1):
        col_len = df[col].astype(str).apply(lambda x: len(x.encode("gbk"))).max()
        col_len = max(len(col.encode("gbk")), col_len) + 1
        worksheet.column_dimensions[get_column_letter(i)].width = col_len
        
        
def render_excel(excel_name, sheet_name=None, suffix="", font="楷体", fontsize=10, theme_color="2639E9"):
    """
    对 excel 文件进行格式渲染
    标题行填充颜色、文本白色楷体加粗
    文本楷体黑色、白色单元格填充
    最外层边框加粗、内层单元格边框常规
    :params excel_name: 需要调整格式的excel文件
    :params sheet_name: 需要调整格式的sheet名称, 可以传入 list 或 str, 不传默认渲染全部 sheet
    :params suffix: 调整后的excel文件保存名称, 不传默认替换文件, 传入字符串会拼在 excel 文件名后
    :params font: 渲染后 excel 的字体, 默认楷体
    :params fontsize: 渲染后 excel 文件的字体大小, 默认 10
    :params theme_color: 渲染后 excel 的主题颜色, 默认 #2639E9
    """
    workbook = load_workbook(excel_name)
    
    if sheet_name and isinstance(worksheet, str):
        sheet_names = [sheet_name]
    else:
        sheet_names = workbook.get_sheet_names()
    
    for sheet_name in sheet_names:
        worksheet = workbook.get_sheet_by_name(sheet_name)
        
        sides = [
            Side(border_style="medium", color=theme_color),
            Side(border_style="thin", color=theme_color),
        ]
        
        for row_index, row in enumerate(worksheet.rows):
            if row_index == 0:
                for col_index, cell in enumerate(row):
                    cell.font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
                    cell.fill = PatternFill(fill_type="solid", start_color=theme_color)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    if col_index == 0:
                        cell.border = Border(left=sides[0], right=sides[1], top=sides[0], bottom=sides[0])
                    elif col_index == len(row) - 1:
                        cell.border = Border(left=sides[1], right=sides[0], top=sides[0], bottom=sides[0])
                    else:
                        cell.border = Border(left=sides[1], right=sides[1], top=sides[0], bottom=sides[0])
            else:
                for col_index, cell in enumerate(row):
                    cell.font = Font(size=fontsize, name=font, color="000000")
                    cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF")
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    if worksheet.max_row == 2:
                        if col_index == 0:
                            cell.border = Border(left=sides[0], right=sides[1], top=sides[0], bottom=sides[0])
                        elif col_index == len(row) - 1:
                            cell.border = Border(left=sides[1], right=sides[0], top=sides[0], bottom=sides[0])
                        else:
                            cell.border = Border(left=sides[1], right=sides[1], top=sides[0], bottom=sides[0])
                    else:
                        if row_index == 1:
                            if col_index == 0:
                                cell.border = Border(left=sides[0], right=sides[1], top=sides[0], bottom=sides[1])
                            elif col_index == len(row) - 1:
                                cell.border = Border(left=sides[1], right=sides[0], top=sides[0], bottom=sides[1])
                            else:
                                cell.border = Border(left=sides[1], right=sides[1], top=sides[0], bottom=sides[1])
                        elif row_index == worksheet.max_row - 1:
                            if col_index == 0:
                                cell.border = Border(left=sides[0], right=sides[1], top=sides[1], bottom=sides[0])
                            elif col_index == len(row) - 1:
                                cell.border = Border(left=sides[1], right=sides[0], top=sides[1], bottom=sides[0])
                            else:
                                cell.border = Border(left=sides[1], right=sides[1], top=sides[1], bottom=sides[0])
                        else:
                            if col_index == 0:
                                cell.border = Border(left=sides[0], right=sides[1], top=sides[1], bottom=sides[1])
                            elif col_index == len(row) - 1:
                                cell.border = Border(left=sides[1], right=sides[0], top=sides[1], bottom=sides[1])
                            else:
                                cell.border = Border(left=sides[1], right=sides[1], top=sides[1], bottom=sides[1])
                                
    excel_name_suffix = excel_name.split(".")[-1]
    workbook.save(excel_name.replace(excel_name_suffix, f"{suffix}{excel_name_suffix}"))
    workbook.close()
    
    
def excel_writer(excel_name, dataframe, sheet_name="Sheet1", index=False, **kwargs):
    with pd.ExcelWriter(excel_name, engine="openpyxl") as writer:
        save_excel(dataframe, writer, sheet_name, index=index)
    render_excel(excel_name, **kwargs)
    
    
if __name__ == "__main__":
    excel_writer("变量清单明细.xlsx", filter_info, sheet_name="附1_变量清单明细")
